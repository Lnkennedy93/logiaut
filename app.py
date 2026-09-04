import streamlit as st
import pandas as pd
import pypdf
import pdfplumber
import urllib.request
import logging
import datetime
from datetime import timezone, timedelta
import re
import os
import io
import base64
import json
from xml.sax.saxutils import escape
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from google.oauth2 import service_account
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from supabase import create_client, Client

# ---------------------------------------------------------
# Configuración inicial de Streamlit
# ---------------------------------------------------------
st.set_page_config(
    page_title="FlowShift",
    page_icon="🔄",
    layout="wide"
)

st.markdown(
    """
    <html lang="es" class="notranslate" translate="no">
    <head>
        <meta name="google" content="notranslate" />
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <style>
        .stDataFrame, .stTable { width: 100% !important; overflow-x: auto !important; }
        div[data-testid="stDataFrame"] div, td, th {
            white-space: normal !important;
            overflow-wrap: anywhere !important;
            word-wrap: break-word !important;
        }
        .product-card {
            width: 100%;
            box-sizing: border-box;
            padding: 1rem;
            margin: 0.75rem 0 1rem;
            border: 1px solid #d9e2f3;
            border-left: 5px solid #1F3864;
            border-radius: 8px;
            background: #f7faff;
            overflow-wrap: anywhere;
            word-break: break-word;
        }
        .product-card p { margin: 0.35rem 0; line-height: 1.5; }
        .empaque-box {
            width: 100%;
            box-sizing: border-box;
            padding: 1rem;
            margin: 0.75rem 0 1rem;
            border: 1px solid #d9e2f3;
            border-radius: 8px;
            background: #f8f9fa;
        }
        @media (max-width: 640px) {
            .block-container { padding: 1rem 0.75rem 2rem; }
            .product-card { padding: 0.8rem; }
        }
    </style>
    """,
    unsafe_allow_html=True
)

# ---------------------------------------------------------
# Parámetros y Constantes
# ---------------------------------------------------------
GOOGLE_SHEET_XLSX_URL = "https://docs.google.com/spreadsheets/d/1aTlmA6JBldTX3zN-djDjWA5HEAExTPcdNhJsPJL9Kgo/export?format=xlsx"
DRIVE_FOLDER_ID = "1Amwy8_uQgo6X0VS2DXH028Ep80BMi4rP"
CEDULA_DEV_CORRECTA = "1073513861"
BD_LOCAL_PATH = "BD_Pesos.xlsx"
MASTER_CSV_PATH = "registro_entregas.csv"
LOGO_PATH = "logo.png" if os.path.exists("logo.png") else ("download.png" if os.path.exists("download.png") else None)

# --- Logging ---
LOG_FILENAME = "log_ejecucion.txt"

logging.basicConfig(
    filename=LOG_FILENAME,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    encoding="utf-8"
)

def registrar_log(mensaje, tipo="INFO"):
    if tipo == "INFO":
        logging.info(mensaje)
    elif tipo == "WARNING":
        logging.warning(mensaje)
    elif tipo == "ERROR":
        logging.error(mensaje)

registrar_log("--- Sesión iniciada ---")

# ---------------------------------------------------------
# Conexión Inteligente Google Drive (Híbrida: Local/Nube)
# ---------------------------------------------------------
def get_google_creds():
    scope = ["https://www.googleapis.com/auth/drive"]
    if "google" in st.secrets:
        creds_dict = dict(st.secrets["google"])
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        return service_account.Credentials.from_service_account_info(creds_dict, scopes=scope)
    elif os.path.exists('credentials.json'):
        return ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    return None

def descargar_pdf_desde_drive(folder_id, numero_entrega):
    try:
        creds = get_google_creds()
        if not creds:
            registrar_log("No se encontraron credenciales de Google Drive.", "ERROR")
            return None
            
        service = build('drive', 'v3', credentials=creds)

        # Consulta robusta buscando tanto dentro del ID de la carpeta como de forma global por si acaso
        query = f"('{folder_id}' in parents or name contains '{numero_entrega}') and mimeType='application/pdf' and trashed=false"
        results = service.files().list(
            q=query, 
            pageSize=5, 
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()
        files = results.get('files', [])

        # Filtramos para asegurar que coincida con el número de entrega exacto
        matched_file = None
        for f in files:
            if str(numero_entrega) in f['name']:
                matched_file = f
                break

        if not matched_file and files:
            matched_file = files[0] # Tomamos el primero como respaldo si contiene coincidencia parcial

        if not matched_file:
            registrar_log(f"No se encontró PDF en Drive para la entrega: {numero_entrega}", "WARNING")
            return None

        file_id = matched_file['id']
        file_name = matched_file['name']

        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()

        fh.seek(0)
        registrar_log(f"PDF descargado exitosamente desde Drive: {file_name}")
        return fh

    except Exception as e:
        registrar_log(f"Error al descargar PDF de Drive para {numero_entrega}: {e}", "ERROR")
        return None

# ---------------------------------------------------------
# Conexión a Supabase (Autenticación y Licencias)
# ---------------------------------------------------------
@st.cache_resource
def init_supabase():
    url = st.secrets["supabase"]["url"]
    key = st.secrets["supabase"]["key"]
    return create_client(url, key)

supabase: Client = init_supabase()

if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario_actual" not in st.session_state:
    st.session_state.usuario_actual = None
if "empresa_actual" not in st.session_state:
    st.session_state.empresa_actual = None
if "rol_actual" not in st.session_state:
    st.session_state.rol_actual = "LOGISTICA"

def validar_en_supabase(correo, password):
    try:
        response = supabase.table("usuarios_licencias") \
            .select("*") \
            .eq("correo", correo.strip()) \
            .eq("password", password.strip()) \
            .execute()
        
        data = response.data
        if data and len(data) > 0:
            usuario = data[0]
            if usuario.get("activo") == True:
                return usuario
        return None
    except Exception as e:
        st.error(f"Error de conexión con la base de datos: {e}")
        return None

def registrar_envio_en_supabase(saved_data, df_resumen, observaciones_texto):
    try:
        total_kg = float(df_resumen["Peso Total (KG)"].sum())
        total_docs = len(set(df_resumen["Entrega"]))
        destinos_str = ", ".join(sorted(set(df_resumen["Destino"].astype(str))))

        payload = {
            "usuario": st.session_state.get("usuario_actual", "Desconocido"),
            "peso_total_kg": total_kg,
            "total_documentos": total_docs,
            "conductor_nombre": saved_data.get("d_nombre", ""),
            "conductor_cedula": saved_data.get("d_cedula", ""),
            "vehiculo_placa": saved_data.get("d_placa", ""),
            "empresa_transporte": saved_data.get("d_transp", ""),
            "destinos": destinos_str,
            "observaciones": observaciones_texto,
            "items_json": json.loads(df_resumen.to_json(orient="records")),
            "saved_data_json": saved_data
        }
        response = supabase.table("historial_envios").insert(payload).execute()
        if response.data:
            registro = response.data[0]
            return registro.get("consecutivo", registro.get("id"))
        return None
    except Exception as e:
        registrar_log(f"Error registrando envío en Supabase: {e}", "ERROR")
        return None

def registrar_envio_local_en_supabase(saved_data, df_resumen, observaciones_texto):
    try:
        payload = {
            "usuario": st.session_state.get("usuario_actual", "Desconocido"),
            "peso_total_kg": float(df_resumen["Peso Total (KG)"].sum()),
            "total_documentos": len(set(df_resumen["Entrega"])),
            "conductor_nombre": saved_data.get("d_nombre", ""),
            "conductor_cedula": saved_data.get("d_cedula", ""),
            "vehiculo_placa": saved_data.get("d_placa", ""),
            "empresa_transporte": saved_data.get("d_transp", ""),
            "destinos": ", ".join(sorted(set(df_resumen["Destino"].astype(str)))),
            "observaciones": observaciones_texto,
            "items_json": json.loads(df_resumen.to_json(orient="records")),
            "saved_data_json": saved_data
        }
        response = supabase.table("historial_envios_local").insert(payload).execute()
        if response.data:
            registro = response.data[0]
            return registro.get("consecutivo", registro.get("id"))
        return None
    except Exception as e:
        registrar_log(f"Error registrando envío local en Supabase: {e}", "ERROR")
        return None

def formatear_fecha_colombia(fecha_raw):
    if not fecha_raw:
        return "N/A"
    try:
        fecha_texto = str(fecha_raw).replace("Z", "+00:00")
        fecha = datetime.datetime.fromisoformat(fecha_texto)
        if fecha.tzinfo is None:
            fecha = fecha.replace(tzinfo=timezone.utc)
        hora_colombia = timezone(timedelta(hours=-5))
        return fecha.astimezone(hora_colombia).strftime("%d-%m-%Y %H:%M:%S")
    except (TypeError, ValueError):
        return str(fecha_raw)[:19].replace("T", " ") or "N/A"

def pantalla_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.subheader("🔄 FlowShift")
        st.caption("La evolución inteligente de tu gestión administrativa.")
        with st.form("form_login"):
            correo = st.text_input("Correo electrónico")
            password = st.text_input("Contraseña", type="password")
            boton_login = st.form_submit_button("Iniciar Sesión", use_container_width=True)

            if boton_login:
                user_data = validar_en_supabase(correo, password)
                if user_data:
                    st.session_state.autenticado = True
                    st.session_state.usuario_actual = user_data["correo"]
                    st.session_state.empresa_actual = user_data["empresa"]
                    st.session_state.rol_actual = str(user_data.get("rol", user_data.get("role", "LOGISTICA"))).strip().upper()
                    st.success("¡Licencia verificada con éxito!")
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas o licencia suspendida/inactiva.")
        
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            """
            <div style="text-align: center; color: gray; font-size: 14px;">
                <p>Pensado para facilitar tu trabajo</p>
                <p><b>Elaborado por Liontech</b></p>
            </div>
            """,
            unsafe_allow_html=True
        )

if not st.session_state.autenticado:
    pantalla_login()
    st.stop()

# ---------------------------------------------------------
# Capa de Datos Inteligente (Patrón Clean Core - SAP BTP / Google)
# ---------------------------------------------------------
def get_product_data_from_source(codigo_input, df_bd):
    env_mode = st.session_state.get("env_mode", "DEV (Google)")
    
    if env_mode == "PROD (SAP)":
        registrar_log(f"[SAP BTP ODATA MOCK] Consumiendo API_PRODUCT_SRV para material: {codigo_input}", "INFO")
        
        if df_bd is not None:
            match = df_bd[
                (df_bd['Codigo'].astype(str).str.contains(codigo_input, case=False, na=False)) |
                (df_bd['Descripcion'].astype(str).str.contains(codigo_input, case=False, na=False))
            ]
            if not match.empty:
                item_sap = match.iloc[0].copy()
                item_sap['Descripcion'] = f"[SAP S/4HANA] {item_sap['Descripcion']}"
                return item_sap
                
        return pd.Series({
            "Codigo": codigo_input.strip().upper(), 
            "Descripcion": f"[SAP S/4HANA] MATERIAL NO ENCONTRADO EN MAESTRO ({codigo_input})", 
            "Peso_KG": 0.0
        })
            
    else:
        if df_bd is not None:
            match = df_bd[
                (df_bd['Codigo'].astype(str).str.contains(codigo_input, case=False, na=False)) |
                (df_bd['Descripcion'].astype(str).str.contains(codigo_input, case=False, na=False))
            ]
            if not match.empty:
                return match.iloc[0]
        return None

@st.cache_data(ttl=1800, show_spinner=False)
def fetch_google_sheet_database(custom_url=None):
    target_url = custom_url or GOOGLE_SHEET_XLSX_URL
    if 'format=csv' in target_url:
        target_url = target_url.replace('format=csv', 'format=xlsx')

    registrar_log(f"Iniciando sincronización completa desde Google Sheets: {target_url}")

    try:
        req = urllib.request.Request(
            target_url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        )
        
        with urllib.request.urlopen(req, timeout=30) as response:
            content = response.read()

        hojas_dict = pd.read_excel(io.BytesIO(content), sheet_name=None, engine='openpyxl')
        list_dfs = []

        for nombre_hoja, df_hoja in hojas_dict.items():
            if df_hoja.empty:
                continue

            df_h = df_hoja.copy()
            
            idx_header = 0
            for idx, row in df_h.iterrows():
                row_str = " ".join([str(val).lower() for val in row.values if pd.notna(val)])
                if any(k in row_str for k in ['código', 'codigo', 'artículo', 'articulo', 'material']) and any(k in row_str for k in ['peso', 'kg']):
                    idx_header = idx
                    break

            if idx_header > 0:
                df_h = pd.read_excel(io.BytesIO(content), sheet_name=nombre_hoja, skiprows=idx_header+1, engine='openpyxl')

            df_h.columns = df_h.columns.astype(str).str.strip()

            renames = {}
            for col in df_h.columns:
                c_low = col.lower()
                if any(k in c_low for k in ['código', 'codigo', 'artículo', 'articulo', 'número', 'numero', 'material']) and 'Codigo' not in renames.values():
                    renames[col] = 'Codigo'
                elif 'desc' in c_low and 'Descripcion' not in renames.values():
                    renames[col] = 'Descripcion'
                elif any(k in c_low for k in ['peso', 'kg']) and 'Peso_KG' not in renames.values():
                    renames[col] = 'Peso_KG'

            df_h = df_h.rename(columns=renames)
            cols_deseadas = [c for c in ['Codigo', 'Descripcion', 'Peso_KG'] if c in df_h.columns]
            
            if 'Codigo' in cols_deseadas and 'Peso_KG' in cols_deseadas:
                df_h = df_h[cols_deseadas].dropna(subset=['Codigo'])
                df_h['PESTAÑA_BD'] = nombre_hoja
                list_dfs.append(df_h)

        if list_dfs:
            df_final = pd.concat(list_dfs, ignore_index=True)
            df_final['Codigo'] = (
                df_final['Codigo']
                .astype(str)
                .str.strip()
                .str.replace(r'\.0$', '', regex=True)
                .str.lstrip('0')
            )
            registrar_log(f"Sincronización exitosa. Total: {len(df_final)} productos en {len(list_dfs)} pestañas.")
            return df_final
        else:
            return None

    except Exception as err:
        registrar_log(f"Error al sincronizar con Google Sheets: {err}", 'ERROR')
        return None

@st.cache_data(ttl=3600, show_spinner=False)
def cargar_bd_local(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        return None
    try:
        hojas_dict = pd.read_excel(ruta_archivo, sheet_name=None, engine='openpyxl')
        list_dfs = []

        for nombre_hoja, df_hoja in hojas_dict.items():
            if not df_hoja.empty:
                df_h = df_hoja.copy()
                df_h.columns = df_h.columns.astype(str).str.strip()

                renames = {}
                for col in df_h.columns:
                    c_low = col.lower()
                    if any(k in c_low for k in ['código', 'codigo', 'artículo', 'articulo', 'número', 'numero', 'material']) and 'Codigo' not in renames.values():
                        renames[col] = 'Codigo'
                    elif 'desc' in c_low and 'Descripcion' not in renames.values():
                        renames[col] = 'Descripcion'
                    elif any(k in c_low for k in ['peso', 'kg']) and 'Peso_KG' not in renames.values():
                        renames[col] = 'Peso_KG'

                df_h = df_h.rename(columns=renames)
                cols_deseadas = [c for c in ['Codigo', 'Descripcion', 'Peso_KG'] if c in df_h.columns]
                
                if 'Codigo' in cols_deseadas and 'Peso_KG' in cols_deseadas:
                    df_h = df_h[cols_deseadas].dropna(subset=['Codigo'])
                    df_h['PESTAÑA_BD'] = nombre_hoja
                    list_dfs.append(df_h)

        if list_dfs:
            df_final = pd.concat(list_dfs, ignore_index=True)
            df_final['Codigo'] = (
                df_final['Codigo']
                .astype(str)
                .str.strip()
                .str.replace(r'\.0$', '', regex=True)
                .str.lstrip('0')
            )
            return df_final
        return None
    except Exception as e:
        registrar_log(f"Error al leer BD local: {e}", "ERROR")
        return None

# ---------------------------------------------------------
# Generadores Excel y PDF (GID-F-010)
# ---------------------------------------------------------
def generar_excel_bytes(df_resumen, total_docs, total_kg, total_ton, driver_info=None, dest_info=None, elaborado_info=None, empaques_info=None, consecutivo_num=None):
    if driver_info is None: driver_info = {}
    if dest_info is None: dest_info = {}
    if elaborado_info is None: elaborado_info = {}
    if empaques_info is None: empaques_info = "Ninguno especificado"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RELACION DE ENVIO"
    ws.views.sheetView[0].showGridLines = True

    font_brand_bold_10 = Font(name="Arial", size=10, bold=True, color="1F3864")
    font_brand_bold_14 = Font(name="Arial", size=14, bold=True, color="1F3864")
    font_arial_bold_10 = Font(name="Arial", size=10, bold=True)
    font_arial_bold_11 = Font(name="Arial", size=11, bold=True)
    font_arial_bold_9 = Font(name="Arial", size=9, bold=True)
    font_arial_9 = Font(name="Arial", size=9, bold=False)
    font_calibri_9_bold = Font(name="Calibri", size=9, bold=True)
    font_calibri_8 = Font(name="Calibri", size=8, italic=True)

    inline_font_bold = InlineFont(b=True, rFont="Calibri", sz=9)
    inline_font_red = InlineFont(b=True, color="FF0000", rFont="Calibri", sz=9)

    fill_header_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_section_gray = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    ws.merge_cells("A1:B5")
    cell_a1 = ws["A1"]
    cell_a1.value = "Sistema de Gestión Integral\nTUVACOL S.A."
    cell_a1.font = font_brand_bold_10
    cell_a1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if LOGO_PATH and os.path.exists(LOGO_PATH):
        try:
            img = OpenpyxlImage(LOGO_PATH)
            img.width = 110
            img.height = 55
            ws.add_image(img, "A1")
        except Exception as ie:
            registrar_log(f"No se pudo cargar la imagen en Excel: {ie}", "WARNING")

    ws.merge_cells("C1:E5")
    cell_c1 = ws["C1"]
    cell_c1.value = "RELACION DE ENVIO DE MERCANCIAS"
    cell_c1.font = font_brand_bold_14
    cell_c1.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("F1:G1")
    ws["F1"] = "Versión: 05"
    ws.merge_cells("F2:G2")
    ws["F2"] = "Vigente desde: 12-02-2021"
    ws.merge_cells("F3:G3")
    ws["F3"] = "Codigo: GID-F-010"
    ws.merge_cells("F4:G4")
    ws["F4"] = "Elaborado por: Comité SGI"
    ws.merge_cells("F5:G5")
    ws["F5"] = "Revisado y Aprobado por: Comité SGI"

    for r in range(1, 6):
        for c in range(6, 8):
            ws.cell(r, c).font = font_arial_9
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

    for r in range(1, 6):
        for c in range(1, 8):
            ws.cell(r, c).border = thin_border

    ws["A7"] = "Responsable / Cargo:"
    ws["A7"].font = font_arial_bold_10

    ws.merge_cells("C7:D7")
    ws["C7"] = "LOGÍSTICA SEDE FUNZA."
    ws["C7"].font = font_arial_bold_10
    ws["C7"].alignment = Alignment(horizontal="center")

    num_relacion = f"No.{str(consecutivo_num).zfill(8)}" if consecutivo_num is not None else "No.PENDIENTE"
    ws["E7"] = num_relacion
    ws["E7"].font = font_arial_bold_11
    ws["E7"].alignment = Alignment(horizontal="center")

    ws["F7"] = "Fecha:"
    ws["F7"].font = font_arial_bold_10
    ws["G7"] = datetime.datetime.now().strftime("%d-%m-%Y")
    ws["G7"].font = font_arial_bold_10
    ws["G7"].alignment = Alignment(horizontal="center")

    for c in range(1, 8):
        ws.cell(7, c).border = thin_border

    headers = [
        "Unidades enviadas",
        "Descripción del empaque",
        "Descripción de la mercancía",
        "",
        "Código de la mercancía",
        "Peso (Kilos)",
        "Documentos de referencia"
    ]

    ws.merge_cells("C9:D9")
    for col_idx, text in enumerate(headers, 1):
        if col_idx == 4:
            continue
        cell = ws.cell(9, col_idx, text)
        cell.font = font_calibri_9_bold
        cell.fill = fill_header_gray
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.cell(9, 4).border = thin_border
    ws.cell(9, 4).fill = fill_header_gray
    ws.row_dimensions[9].height = 28

    current_row = 10
    tot_und = 0.0
    tot_mts = 0.0
    docs_unicos = []
    filas_por_documento = []
    destino_actual = None

    for idx, row in df_resumen.iterrows():
        cant_val = float(row["Cantidad"])
        desc_str = str(row["Descripción"])
        doc_num = str(row["Entrega"])
        destino = str(row.get("Destino", "DESTINO GENERAL")).strip() or "DESTINO GENERAL"

        if destino != destino_actual:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            ws.cell(current_row, 1, f"DESTINO: {destino}")
            ws.cell(current_row, 1).font = font_arial_bold_10
            ws.cell(current_row, 1).fill = fill_header_gray
            ws.cell(current_row, 1).alignment = Alignment(horizontal="left")
            for col in range(1, 8):
                ws.cell(current_row, col).border = thin_border
                ws.cell(current_row, col).fill = fill_header_gray
            current_row += 1
            destino_actual = destino

        if doc_num and doc_num not in docs_unicos:
            docs_unicos.append(doc_num)

        es_promediado = (cant_val % 1 != 0)

        if "TUBERIA" in desc_str.upper():
            empaque_tag = "TUBERIA"
            unidad_str = f"{cant_val:,.0f} MTS"
            tot_mts += cant_val
        else:
            empaque_tag = "ACCESORIOS"
            unidad_str = f"{cant_val:,.0f} UND"
            tot_und += cant_val

        cell_unidades = ws.cell(current_row, 1)
        cell_unidades.alignment = Alignment(horizontal="center")

        if es_promediado and "TUBERIA" in desc_str.upper():
            rich_unidades = CellRichText(
                TextBlock(inline_font_bold, unidad_str),
                TextBlock(inline_font_red, " Promedio")
            )
            cell_unidades.value = rich_unidades
        else:
            cell_unidades.value = unidad_str

        ws.cell(current_row, 2, empaque_tag).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        ws.cell(current_row, 3, desc_str).alignment = Alignment(horizontal="left")
        ws.cell(current_row, 5, str(row["Código"])).alignment = Alignment(horizontal="center")

        p_tot = float(row.get("Peso Total (KG)", 0.0))
        filas_por_documento.append((current_row, doc_num))
        ws.cell(current_row, 6, p_tot).alignment = Alignment(horizontal="right")
        ws.cell(current_row, 6).number_format = "#,##0.00"
        ws.cell(current_row, 7, doc_num).alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 8):
            c_cell = ws.cell(current_row, col)
            if not (col == 1 and es_promediado and "TUBERIA" in desc_str.upper()):
                c_cell.font = font_calibri_9_bold
            c_cell.border = thin_border
        current_row += 1

    i = 0
    while i < len(filas_por_documento):
        r_first, doc_val = filas_por_documento[i]
        j = i + 1
        while j < len(filas_por_documento) and filas_por_documento[j][1] == doc_val:
            j += 1
        r_last = filas_por_documento[j - 1][0]
        if r_last > r_first:
            ws.merge_cells(start_row=r_first, start_column=7, end_row=r_last, end_column=7)
            cell_m = ws.cell(row=r_first, column=7)
            cell_m.value = doc_val
            cell_m.alignment = Alignment(horizontal="center", vertical="center")
        i = j

    docs_str = ", ".join(docs_unicos)
    obs_texto = f"Observaciones: TOTAL: {tot_und:,.0f} UND, {tot_mts:,.0f} MTS | DOCS: {docs_str} | Empaques: {empaques_info}"

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(current_row, 1, obs_texto).font = font_calibri_9_bold

    ws.cell(current_row, 5, "Peso Total:").font = font_calibri_9_bold
    ws.cell(current_row, 5).fill = fill_header_gray
    ws.cell(current_row, 5).alignment = Alignment(horizontal="right")

    ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
    cell_tot = ws.cell(current_row, 6, f"=SUM(F10:F{current_row-1})")
    cell_tot.font = font_arial_bold_11
    cell_tot.alignment = Alignment(horizontal="center")
    cell_tot.number_format = "#,##0.00"

    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 2

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    c_rem = ws.cell(current_row, 1, "Datos del Remitente")
    c_rem.font = font_arial_bold_10
    c_rem.fill = fill_header_gray

    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=7)
    c_dest = ws.cell(current_row, 4, "Datos del Destinatario")
    c_dest.font = font_arial_bold_10
    c_dest.fill = fill_header_gray

    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.cell(current_row, 1, "Nombre:").font = font_arial_9
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.cell(current_row, 2, "TUVACOL SEDE FUNZA").font = font_arial_bold_10

    ws.cell(current_row, 4, "Nombre:").font = font_arial_9
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    ws.cell(current_row, 5, dest_info.get("nombre", "DESPACHOS CLIENTE / REMISIÓN CONSOLIDADA")).font = font_arial_bold_10

    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.cell(current_row, 1, "Dirección y Teléfono:").font = font_arial_9
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.cell(current_row, 2, "KM 3,5 VIA FUNZA SIBERIA PARQUE INDUSTRIAL SAN JOSE BODEGA 1 MANZANA B // TEL: 823 77 79").font = Font(name="Arial", size=8)

    ws.cell(current_row, 4, "Dirección y Teléfono:").font = font_arial_9
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    ws.cell(current_row, 5, dest_info.get("direccion", "DIRECCIÓN DE DESTINO REGISTRADA EN REMISIONES")).font = Font(name="Arial", size=8)

    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    c_cond = ws.cell(current_row, 1, "Datos del Conductor y Vehiculo")
    c_cond.font = font_arial_bold_10
    c_cond.fill = fill_section_gray
    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.cell(current_row, 1, "Nombre:").font = font_arial_bold_9
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.cell(current_row, 2, driver_info.get("nombre", "")).font = font_arial_9
    ws.cell(current_row, 4, "N. de Placa:").font = font_arial_bold_9
    ws.cell(current_row, 5, driver_info.get("placa", "")).font = font_arial_9
    ws.cell(current_row, 6, "Modelo/Marca:").font = font_arial_bold_9
    ws.cell(current_row, 7, driver_info.get("marca", "")).font = font_arial_9
    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.cell(current_row, 1, "Cedula No.:").font = font_arial_bold_9
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.cell(current_row, 2, driver_info.get("cedula", "")).font = font_arial_9
    ws.cell(current_row, 4, "Transportadora:").font = font_arial_bold_9
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    ws.cell(current_row, 5, driver_info.get("transportadora", "")).font = font_arial_9
    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 1

    ws.cell(current_row, 1, "N. de Celular:").font = font_arial_bold_9
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.cell(current_row, 2, driver_info.get("celular", "")).font = font_arial_9
    for col in range(1, 8):
        ws.cell(current_row, col).border = thin_border
    current_row += 2

    elaborado_nombre = elaborado_info.get("nombre", "LOGÍSTICA TUVACOL S.A.")
    ws.cell(current_row, 1, f"Elaborado por:\n{elaborado_nombre}").font = font_arial_9
    ws.cell(current_row, 4, "Firma y cédula del Conductor:").font = font_arial_9
    current_row += 2
    ws.cell(current_row, 1, "Recibe y acepta (Firma y Sello):").font = font_arial_9
    current_row += 2

    ws.cell(current_row, 1, "*El remitente declara que se envía exactamente la mercancía relacionada en este documento.").font = font_calibri_8
    current_row += 1
    ws.cell(current_row, 1, "*El conductor declara que ha recibido a satisfacción lo que en este documento se relaciona incluyendo cantidades, peso y empaques.").font = font_calibri_8

    col_widths = {'A': 20, 'B': 16, 'C': 30, 'D': 35, 'E': 18, 'F': 14, 'G': 22}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_pdf_bytes(df_resumen, total_docs, total_kg, total_ton, driver_info=None, dest_info=None, elaborado_info=None, empaques_info=None, consecutivo_num=None, es_local=False):
    if driver_info is None: driver_info = {}
    if dest_info is None: dest_info = {}
    if elaborado_info is None: elaborado_info = {}
    if empaques_info is None: empaques_info = "Ninguno especificado"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15
    )
    elements = []

    styles = getSampleStyleSheet()
    normal_style = ParagraphStyle('Normal8', parent=styles['Normal'], fontSize=7.5, leading=8)
    bold_style = ParagraphStyle('Bold8', parent=normal_style, fontName='Helvetica-Bold')
    
    brand_title_style = ParagraphStyle(
        'BrandTitle10', parent=styles['Normal'], 
        fontName='Helvetica-Bold', fontSize=10, leading=10, alignment=1,
        textColor=colors.HexColor('#1F3864')
    )

    if LOGO_PATH and os.path.exists(LOGO_PATH):
        try:
            img_element = ReportLabImage(LOGO_PATH, width=110, height=45)
            sgi_cell = img_element
        except Exception:
            sgi_cell = Paragraph("<b>Sistema de Gestión Integral</b><br/><font size=9><b>TUVACOL S.A.</b></font>", brand_title_style)
    else:
        sgi_cell = Paragraph("<b>Sistema de Gestión Integral</b><br/><font size=9><b>TUVACOL S.A.</b></font>", brand_title_style)

    h_data = [
        [
            sgi_cell,
            Paragraph(
                "<b>RELACION DESPACHOS LOCALES</b>" if es_local else "<b>RELACION DE ENVIO DE MERCANCIAS</b>",
                brand_title_style
            ),
            Paragraph("Versión: 05<br/>Vigente desde: 12-02-2021<br/><b>Codigo: GID-F-010</b><br/>Elaborado por: Comité SGI", normal_style)
        ]
    ]
    t_header = Table(h_data, colWidths=[150, 270, 150])
    t_header.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    elements.append(t_header)
    elements.append(Spacer(1, 2))

    fecha_act = datetime.datetime.now().strftime("%d-%m-%Y")
    num_rel = f"No.{str(consecutivo_num).zfill(8)}" if consecutivo_num is not None else "No.PENDIENTE"
    m_data = [
        [
            Paragraph("<b>Responsable / Cargo:</b>", normal_style),
            Paragraph("<b>LOGÍSTICA SEDE FUNZA.</b>", bold_style),
            Paragraph(f"<b>{num_rel}</b>", brand_title_style),
            Paragraph("<b>Fecha:</b>", normal_style),
            Paragraph(f"<b>{fecha_act}</b>", bold_style)
        ]
    ]
    t_meta = Table(m_data, colWidths=[100, 230, 100, 50, 90])
    t_meta.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (0,0), colors.HexColor('#D9D9D9')),
        ('BACKGROUND', (3,0), (3,0), colors.HexColor('#D9D9D9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (2,0), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 2))
    d_table_data = [
        [
            Paragraph("<b>Unidades enviadas</b>", normal_style),
            Paragraph("<b>Descripción del empaque</b>", normal_style),
            Paragraph("<b>Descripción de la mercancía</b>", normal_style),
            Paragraph("<b>Código de la mercancía</b>", normal_style),
            Paragraph("<b>Peso (Kilos)</b>", normal_style),
            Paragraph("<b>Doc. Ref.</b>", normal_style)
        ]
    ]

    tot_und = 0.0
    tot_mts = 0.0
    docs_unicos = []
    filas_por_documento = []
    filas_por_destino = []
    destino_actual = None

    def agregar_fila_material(row):
        nonlocal tot_und, tot_mts
        p_tot = float(row.get("Peso Total (KG)", 0.0))
        doc_num = str(row["Entrega"])
        if doc_num and doc_num not in docs_unicos:
            docs_unicos.append(doc_num)

        cant_val = float(row["Cantidad"])
        desc_str = str(row["Descripción"])
        es_promediado = (cant_val % 1 != 0)

        if "TUBERIA" in desc_str.upper():
            empaque_tag = "TUBERIA"
            unidad_str = f"{cant_val:,.0f} MTS"
            tot_mts += cant_val
        else:
            empaque_tag = "ACCESORIOS"
            unidad_str = f"{cant_val:,.0f} UND"
            tot_und += cant_val

        filas_por_documento.append((len(d_table_data), doc_num))
        d_table_data.append([
            Paragraph(f"<b>{unidad_str}</b>", normal_style),
            Paragraph(empaque_tag, normal_style),
            Paragraph(desc_str, normal_style),
            Paragraph(f"<b>{row['Código']}</b>", normal_style),
            Paragraph(f"<b>{p_tot:,.2f}</b>", normal_style),
            Paragraph(f"<b>{doc_num}</b>", normal_style)
        ])

    entregas = df_resumen["Entrega"].astype(str)
    df_edm = df_resumen[entregas.str.upper().str.startswith("EDM", na=False)]
    df_otros = df_resumen[~entregas.str.upper().str.startswith("EDM", na=False)]

    if es_local:
        for entrega, group_doc in df_resumen.groupby("Entrega", sort=False):
            cliente_val = group_doc["Cliente"].iloc[0] if "Cliente" in group_doc.columns else "CLIENTE GENERAL"
            peso_doc_total = float(group_doc["Peso Total (KG)"].sum())
            filas_por_destino.append(len(d_table_data))
            d_table_data.append([
                Paragraph(
                    f"<b>DOCUMENTO: {escape(str(entrega))} | CLIENTE: {escape(str(cliente_val))} | PESO TOTAL DOC: {peso_doc_total:,.2f} KG</b>",
                    bold_style
                ),
                "", "", "", "", ""
            ])
            for _, row in group_doc.iterrows():
                agregar_fila_material(row)
    elif not df_edm.empty:
        edm_nums = []
        for entrega in df_edm["Entrega"].unique():
            coincidencia = re.search(r"\d+", str(entrega))
            edm_nums.append(coincidencia.group() if coincidencia else str(entrega))
        edm_str = ", ".join(edm_nums)
        filas_por_destino.append(len(d_table_data))
        d_table_data.append([
            Paragraph(f"<b>ENTREGA DE MERCANCIA: {escape(edm_str)}</b>", bold_style),
            "", "", "", "", ""
        ])
        for _, row in df_edm.iterrows():
            agregar_fila_material(row)

    if not es_local:
        for destino, group_dest in df_otros.groupby("Destino"):
            filas_por_destino.append(len(d_table_data))
            d_table_data.append([
                Paragraph(f"<b>DESTINO: {escape(str(destino))}</b>", bold_style),
                "", "", "", "", ""
            ])
            for _, row in group_dest.iterrows():
                agregar_fila_material(row)

    docs_str = ", ".join(docs_unicos)
    obs_texto = f"Observaciones: TOTAL: {tot_und:,.0f} UND, {tot_mts:,.0f} MTS | DOCS: {docs_str} | Empaques: {empaques_info}"

    d_table_data.append([
        Paragraph(f"<b>{obs_texto}</b>", normal_style),
        "", "",
        Paragraph("<b>Peso Total:</b>", bold_style),
        Paragraph(f"<b>{total_kg:,.2f}</b>", bold_style),
        Paragraph(f"<b>{total_ton:,.3f} T</b>", bold_style)
    ])
    ultima_fila_datos = len(d_table_data) - 1

    table_styles = [
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9D9D9')),
        ('SPAN', (0, ultima_fila_datos), (2, ultima_fila_datos)),
        ('BACKGROUND', (3, ultima_fila_datos), (3, ultima_fila_datos), colors.HexColor('#D9D9D9')),
        ('BACKGROUND', (5, ultima_fila_datos), (5, ultima_fila_datos), colors.HexColor('#D9D9D9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,1), (1,-2), 'CENTER'),
        ('ALIGN', (5,1), (5,-2), 'CENTER'),
    ]
    for row_index in filas_por_destino:
        table_styles.extend([
            ('SPAN', (0, row_index), (5, row_index)),
            ('BACKGROUND', (0, row_index), (5, row_index), colors.HexColor('#D9D9D9')),
            ('ALIGN', (0, row_index), (5, row_index), 'LEFT'),
        ])
    i = 0
    while i < len(filas_por_documento):
        r_start, doc_val = filas_por_documento[i]
        j = i + 1
        while j < len(filas_por_documento) and filas_por_documento[j][1] == doc_val:
            j += 1
        r_end = filas_por_documento[j - 1][0]
        if r_end > r_start:
            table_styles.append(('SPAN', (5, r_start), (5, r_end)))
        i = j

    t_data = Table(d_table_data, colWidths=[65, 65, 245, 75, 62, 60])
    t_data.setStyle(TableStyle(table_styles))
    elements.append(t_data)
    elements.append(Spacer(1, 2))

    log_data = [
        [Paragraph("<b>Datos del Remitente</b>", bold_style), "", Paragraph("<b>Datos del Destinatario</b>", bold_style), ""],
        [
            Paragraph("<b>Nombre:</b>", normal_style), 
            Paragraph("TUVACOL SEDE FUNZA", bold_style), 
            Paragraph("<b>Nombre:</b>", normal_style), 
            Paragraph(f"<b>{dest_info.get('nombre', 'DESPACHOS CLIENTE / REMISIÓN')}</b>", bold_style)
        ],
        [
            Paragraph("<b>Dirección/Tel:</b>", normal_style), 
            Paragraph("KM 3,5 VIA FUNZA SIBERIA BODEGA 1", normal_style), 
            Paragraph("<b>Dirección/Tel:</b>", normal_style), 
            Paragraph(dest_info.get("direccion", "DIRECCIÓN DE DESTINO REGISTRADA EN REMISIONES"), normal_style)
        ]
    ]
    t_log = Table(log_data, colWidths=[70, 215, 70, 215])
    t_log.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('SPAN', (0,0), (1,0)),
        ('SPAN', (2,0), (3,0)),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9D9D9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    if not es_local:
        elements.append(t_log)
        elements.append(Spacer(1, 2))

    drv_data = [
        [Paragraph("<b>Datos del Conductor y Vehiculo</b>", bold_style), "", "", "", "", ""],
        [
            Paragraph("<b>Nombre:</b>", normal_style), Paragraph(driver_info.get("nombre") or "_______________________", normal_style), 
            Paragraph("<b>Placa:</b>", normal_style), Paragraph(driver_info.get("placa") or "______", normal_style), 
            Paragraph("<b>Modelo/Marca:</b>", normal_style), Paragraph(driver_info.get("marca") or "______", normal_style)
        ],
        [
            Paragraph("<b>Cédula No.:</b>", normal_style), Paragraph(driver_info.get("cedula") or "_______________________", normal_style), 
            Paragraph("<b>Empresa:</b>", normal_style), Paragraph(driver_info.get("transportadora") or "__________________", normal_style), "", ""
        ],
        [
            Paragraph("<b>Celular:</b>", normal_style), Paragraph(driver_info.get("celular") or "_______________________", normal_style), "", "", "", ""
        ]
    ]
    if es_local:
        drv_data = [
            [Paragraph("<b>Datos del Conductor y Vehículo</b>", bold_style), ""],
            [Paragraph("<b>Nombre:</b>", normal_style), Paragraph(driver_info.get("nombre") or "_______________________", normal_style)],
            [Paragraph("<b>Cédula No.:</b>", normal_style), Paragraph(driver_info.get("cedula") or "_______________________", normal_style)],
            [Paragraph("<b>Placa:</b>", normal_style), Paragraph(driver_info.get("placa") or "______", normal_style)]
        ]
        t_drv = Table(drv_data, colWidths=[120, 462])
    else:
        t_drv = Table(drv_data, colWidths=[65, 170, 50, 95, 50, 140])
    estilos_conductor = [
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('SPAN', (0,0), (1,0)),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#BFBFBF')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]
    if not es_local:
        estilos_conductor.extend([
            ('SPAN', (3,2), (5,2)),
            ('SPAN', (1,3), (5,3)),
        ])
    t_drv.setStyle(TableStyle(estilos_conductor))
    elements.append(t_drv)
    elements.append(Spacer(1, 4))

    elab_nombre = elaborado_info.get("nombre", "LOGÍSTICA TUVACOL S.A.")
    sig_data = [
        [Paragraph(f"<b>Elaborado por:</b><br/>{elab_nombre}", normal_style), Paragraph("<b>Firma y Cédula del Conductor:</b>", normal_style), Paragraph("<b>Recibe y Acepta:</b>", normal_style)],
        ["", "", ""],
        [Paragraph("___________________________<br/>LOGÍSTICA TUVACOL S.A.", normal_style), Paragraph("___________________________<br/>C.C.", normal_style), Paragraph("___________________________<br/>CLIENTE / DESTINATARIO", normal_style)]
    ]
    t_sig = Table(sig_data, colWidths=[190, 190, 190])
    t_sig.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    elements.append(t_sig)
    elements.append(Spacer(1, 2))

    disc_style = ParagraphStyle('Disc', parent=normal_style, fontSize=6, leading=7, fontName='Helvetica-Oblique')
    elements.append(Paragraph("* El remitente declara que se envía exactamente la mercancía relacionada en este documento.", disc_style))
    elements.append(Paragraph("* El conductor declara que ha recibido a satisfacción lo que en este documento se relaciona incluyendo cantidades, peso y empaques.", disc_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generar_pdf_local_bytes(df_resumen, saved_data, observaciones_texto, consecutivo_num):
    filas_consolidadas = []
    for entrega, grupo_documento in df_resumen.groupby("Entrega", sort=False):
        cliente = grupo_documento["Cliente"].iloc[0] if "Cliente" in grupo_documento else "CLIENTE GENERAL"
        es_mts = grupo_documento["Descripción"].astype(str).str.upper().str.contains("MTS|TUBERIA", na=False)
        for tipo, grupo_tipo in (("MTS", grupo_documento[es_mts]), ("UND", grupo_documento[~es_mts])):
            if grupo_tipo.empty:
                continue
            filas_consolidadas.append({
                "Entrega": entrega,
                "Cliente": cliente,
                "Código": grupo_tipo["Código"].iloc[0] if len(grupo_tipo) == 1 else "VARIOS",
                "Descripción": "TUBERIA CONSOLIDADA" if tipo == "MTS" else "ACCESORIOS / VARIOS CONSOLIDADOS",
                "Instrucciones": grupo_tipo["Instrucciones"].iloc[0] if "Instrucciones" in grupo_tipo.columns else "SIN INSTRUCCIONES",
                "Cantidad": float(grupo_tipo["Cantidad"].sum()),
                "Peso Unit. (KG)": 0.0,
                "Peso Total (KG)": float(grupo_tipo["Peso Total (KG)"].sum()),
                "Destino": grupo_tipo["Destino"].iloc[0]
            })
    df_resumen = pd.DataFrame(filas_consolidadas)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("LocalTitle", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=11, leading=12, alignment=1)
    cell_style = ParagraphStyle("LocalCell", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=9)
    header_style = ParagraphStyle("LocalHeader", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    gray_style = ParagraphStyle("LocalGray", parent=cell_style, fontName="Helvetica-Bold", fontSize=8.5)

    story = []
    consecutivo = f"No.{str(consecutivo_num).zfill(8)}"
    encabezado = Table([[
        Paragraph("<b>TUVACOL S.A.</b><br/>TUBERIAS Y VALVULAS DE COLOMBIA", title_style),
        Paragraph("<b>RELACION DESPACHOS LOCALES</b>", title_style),
        Paragraph(f"<b>{consecutivo}</b><br/>Fecha: {datetime.date.today():%d-%m-%Y}", cell_style)
    ]], colWidths=[150, 260, 172])
    encabezado.setStyle(TableStyle([("GRID", (0, 0), (2, 0), 1, colors.black), ("VALIGN", (0, 0), (2, 0), "MIDDLE"), ("ALIGN", (0, 0), (2, 0), "CENTER"), ("TOPPADDING", (0, 0), (2, 0), 2), ("BOTTOMPADDING", (0, 0), (2, 0), 2)]))
    story.extend([encabezado, Spacer(1, 2)])

    filas = [[
        Paragraph("Unidades enviadas", header_style),
        Paragraph("Descripción del empaque", header_style),
        Paragraph("Instrucciones de Entrega", header_style),
        Paragraph("Peso (Kilos)", header_style),
        Paragraph("Doc. Ref.", header_style)
    ]]
    estilos = [("GRID", (0, 0), (4, -1), 1, colors.black), ("BACKGROUND", (0, 0), (4, 0), colors.HexColor("#1F3864")), ("VALIGN", (0, 0), (4, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (4, -1), 2), ("BOTTOMPADDING", (0, 0), (4, -1), 2)]
    fila_actual = 1

    for entrega, grupo in df_resumen.groupby("Entrega", sort=False):
        cliente = grupo["Cliente"].iloc[0] if "Cliente" in grupo.columns else "CLIENTE GENERAL"
        peso_documento = float(grupo["Peso Total (KG)"].sum())
        filas.append([Paragraph(f"<b>DOCUMENTO: {escape(str(entrega))} | CLIENTE: {escape(str(cliente))} | PESO TOTAL DOC: {peso_documento:,.2f} KG</b>", gray_style), "", "", "", ""])
        estilos.extend([("SPAN", (0, fila_actual), (4, fila_actual)), ("BACKGROUND", (0, fila_actual), (4, fila_actual), colors.HexColor("#D9D9D9"))])
        fila_actual += 1

        es_mts = grupo["Descripción"].astype(str).str.upper().str.contains("MTS|TUBERIA", na=False)
        for tipo, grupo_tipo in (("MTS", grupo[es_mts]), ("UND", grupo[~es_mts])):
            if grupo_tipo.empty:
                continue
            cantidad = float(grupo_tipo["Cantidad"].sum())
            peso = float(grupo_tipo["Peso Total (KG)"].sum())
            filas.append([
                Paragraph(f"{cantidad:,.0f} {tipo}", cell_style),
                Paragraph("TUBERIA" if tipo == "MTS" else "ACCESORIOS", cell_style),
                Paragraph(str(grupo_tipo["Instrucciones"].iloc[0]) if "Instrucciones" in grupo_tipo.columns else "SIN INSTRUCCIONES", cell_style),
                Paragraph(f"{peso:,.2f}", cell_style),
                Paragraph(str(entrega), cell_style)
            ])
            fila_actual += 1

    tabla = Table(filas, colWidths=[65, 75, 287, 55, 100])
    tabla.setStyle(TableStyle(estilos))
    story.extend([tabla, Spacer(1, 2)])

    peso_total = float(df_resumen["Peso Total (KG)"].sum())
    resumen = Table([[
        Paragraph(f"<b>Observaciones:</b> {escape(observaciones_texto)}", cell_style),
        Paragraph(f"<b>Peso Total:</b><br/>{peso_total:,.2f} KG", cell_style),
        Paragraph(f"<b>{peso_total / 1000:,.3f} T</b>", cell_style)
    ]], colWidths=[392, 90, 100])
    resumen.setStyle(TableStyle([("GRID", (0, 0), (2, 0), 1, colors.black), ("VALIGN", (0, 0), (2, 0), "MIDDLE"), ("TOPPADDING", (0, 0), (2, 0), 2), ("BOTTOMPADDING", (0, 0), (2, 0), 2)]))
    story.extend([resumen, Spacer(1, 2)])

    conductor = Table([[
        Paragraph("<b>Datos del Conductor y Vehículo</b>", cell_style), "", ""],
        [Paragraph(f"<b>Conductor:</b> {escape(str(saved_data.get('d_nombre', '')))}", cell_style), Paragraph(f"<b>Cédula:</b> {escape(str(saved_data.get('d_cedula', '')))}", cell_style), Paragraph(f"<b>Placa:</b> {escape(str(saved_data.get('d_placa', '')))}", cell_style)]
    ], colWidths=[194, 194, 194])
    conductor.setStyle(TableStyle([("GRID", (0, 0), (2, 1), 1, colors.black), ("SPAN", (0, 0), (2, 0)), ("BACKGROUND", (0, 0), (2, 0), colors.HexColor("#EAEAEA")), ("TOPPADDING", (0, 0), (2, 1), 2), ("BOTTOMPADDING", (0, 0), (2, 1), 2)]))
    story.extend([conductor, Spacer(1, 4)])
    firmas = Table([[Paragraph(f"<b>Elaborado por:</b><br/>{escape(str(saved_data.get('elab_nombre', '')))}", cell_style), Paragraph("<b>Firma y Cédula del Conductor:</b><br/><br/>___________________________", cell_style), Paragraph("<b>Recibe y Acepta:</b><br/><br/>___________________________", cell_style)]], colWidths=[194, 194, 194])
    firmas.setStyle(TableStyle([("VALIGN", (0, 0), (2, 0), "TOP"), ("TOPPADDING", (0, 0), (2, 0), 2), ("BOTTOMPADDING", (0, 0), (2, 0), 2)]))
    story.append(firmas)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

# ---------------------------------------------------------
# Identificación de documentos y extracción de materiales
# ---------------------------------------------------------
def analizar_metadatos_documento(pdf_source):
    texto_pagina1 = ""
    try:
        if hasattr(pdf_source, "seek"):
            pdf_source.seek(0)
        with pdfplumber.open(pdf_source) as pdf:
            if pdf.pages:
                pagina = pdf.pages[0]
                texto_pagina1 = pagina.extract_text() or ""
                palabras = pagina.extract_words()
                etiquetas_cliente = [p for p in palabras if p["text"].upper().rstrip(":") == "CLIENTE"]
                if etiquetas_cliente:
                    etiqueta = etiquetas_cliente[0]
                    cliente_words = [
                        p for p in palabras
                        if p["top"] >= etiqueta["bottom"] - 1
                        and p["top"] <= etiqueta["bottom"] + 32
                        and p["x1"] >= etiqueta["x0"] - 5
                        and p["x0"] <= etiqueta["x1"] + 260
                        and p["text"].upper().rstrip(":") not in {
                            "NIT", "DIRECCION", "DIRECCIÓN", "CIUDAD", "TELEFONO",
                            "TELÉFONO", "FECHA", "VENCIMIENTO"
                        }
                    ]
                    cliente_words.sort(key=lambda p: (p["top"], p["x0"]))
                    if cliente_words:
                        cliente_nombre = " ".join(p["text"] for p in cliente_words).strip()
    except Exception as error:
        registrar_log(f"No se pudo identificar los metadatos del documento: {error}", "WARNING")
        return "DOCUMENTO No. S/N", "DESTINO GENERAL", "CLIENTE GENERAL", "SIN INSTRUCCIONES"

    tipos_documento = [
        ("TRANSFERENCIA DE STOCK", "TS"),
        ("ENTREGA DE MERCANCÍA", "EDM"),
        ("ENTREGA DE MERCANCIA", "EDM"),
        ("NOTA CRÉDITO", "NC"),
        ("NOTA CREDITO", "NC"),
        ("NOTA DÉBITO", "ND"),
        ("NOTA DEBITO", "ND"),
        ("REMISIÓN", "REM"),
        ("REMISION", "REM"),
        ("FACTURA", "FACT"),
    ]
    patron_numero = r"(?:NO\.?|N[º°]|NUMERO|NÚMERO)\s*[:#]?\s*(\d{4,10})"
    lineas = [re.sub(r"\s+", " ", linea).strip().upper() for linea in texto_pagina1.splitlines() if linea.strip()]

    tipo_doc = "DOCUMENTO"
    numero_doc = "S/N"
    for linea in lineas:
        for texto_tipo, tipo_normalizado in tipos_documento:
            if texto_tipo not in linea:
                continue
            tipo_doc = tipo_normalizado
            match_numero = re.search(
                rf"{re.escape(texto_tipo)}\s*(?:NÚMERO\s*)?{patron_numero}",
                linea
            )
            if match_numero:
                numero_doc = match_numero.group(1)
                break
        if numero_doc != "S/N":
            break

    texto_upper = " ".join(lineas)
    if tipo_doc == "DOCUMENTO":
        for texto_tipo, tipo_normalizado in tipos_documento:
            if texto_tipo in texto_upper:
                tipo_doc = tipo_normalizado
                break

    if numero_doc == "S/N":
        match_numero = re.search(patron_numero, texto_upper)
        if match_numero:
            numero_doc = match_numero.group(1)
        else:
            match_archivo = re.search(r"\d+", getattr(pdf_source, "name", ""))
            if match_archivo:
                numero_doc = match_archivo.group(0)

    doc_clean = f"{tipo_doc} No. {numero_doc}"

    sucursal_destino = "DESTINO GENERAL"
    match_nit = re.search(r"NIT\s*:\s*([^\n]+)", texto_pagina1, re.IGNORECASE)
    if match_nit:
        sucursal_destino = match_nit.group(1).strip()
    else:
        destinos = ["CARTAGENA", "BARRANQUILLA", "MEDELLÍN", "MEDELLIN", "CALI", "TRÁNSITO", "TRANSITO", "BOGOTÁ", "BOGOTA", "FUNZA"]
        for linea in lineas:
            if any(destino in linea for destino in destinos):
                sucursal_destino = linea
                break

    cliente_nombre = "CLIENTE GENERAL"
    for indice, linea in enumerate(lineas):
        if linea == "CLIENTE" and indice + 1 < len(lineas):
            cliente_nombre = lineas[indice + 1]
            break

    instrucciones_entrega = "SIN INSTRUCCIONES"
    for linea in lineas:
        if linea.startswith(("OBSERVACIÓN:", "OBSERVACION:")):
            instrucciones_entrega = linea.split(":", 1)[1].strip() or "SIN INSTRUCCIONES"
            break

    return doc_clean, sucursal_destino, cliente_nombre, instrucciones_entrega

# ---------------------------------------------------------
# Extractor Inteligente de Materiales
# ---------------------------------------------------------
def extraer_tabla_materiales(pdf_source, nombre_doc="Desconocido"):
    texto_completo = ""

    try:
        if hasattr(pdf_source, 'seek'):
            pdf_source.seek(0)
        with pdfplumber.open(pdf_source) as pdf:
            for page in pdf.pages:
                txt = page.extract_text()
                if txt:
                    texto_completo += txt + "\n"
    except Exception as error:
        registrar_log(f"[{nombre_doc}] Error crítico de lectura de PDF con pdfplumber: {error}", "ERROR")
        st.warning(f"⚠️ El archivo `{nombre_doc}` está corrupto o incompleto en el origen. Se omitirá del cálculo.")
        return []

    lineas = [l.strip() for l in texto_completo.split('\n') if l.strip()]
    items = []
    excluidos = ['806014553', '806014553-6', '6658461', '8237779', '10109294']

    for linea_str in lineas:
        if any(term in linea_str.upper() for term in ['DECLARAMOS', 'GARANTÍA', 'RESPONSABLE', 'MERCANCÍA ENTREGADA', 'IMPRESO POR', 'REALIZADO POR', 'PÁGINA', 'CONDICIÓN DE PAGO', 'DIRECCIÓN DE ENTREGA', 'NIT / CC', 'TELÉFONOS', 'ÍTEM', 'CÓDIGO', 'CANTIDAD', 'DESCRIPCIÓN', 'UNIDAD', 'ORDEN DE COMPRA', 'VENDEDOR', 'FECHA', 'ACOPI', 'YUMBO', 'CARRERA', 'CALLE', 'AVENIDA', 'CLIENTE', 'SUCURSAL', 'HIDROOBRAS']):
            continue
        
        if linea_str.startswith('#'):
            continue

        codigo = None
        cant_val = None
        desc_bruta = None

        match_b = re.search(r'^(?:\d+\s+)?([A-Z0-9\-]{7,12})\s+(\d+(?:[\.,]\d+)?)\s+(?:UND|MTS|MT|Und|Unidad)\b\s*(.+)', linea_str, re.IGNORECASE)
        if match_b:
            codigo = match_b.group(1).upper()
            cant_val = float(match_b.group(2).replace(',', '.'))
            desc_bruta = match_b.group(3).strip()
        else:
            match_a = re.search(r'^(?:\d+\s+)?([A-Z0-9\-]{7,12})\s+(\d+(?:[\.,]\d+)?)\s*(?:UND|MTS|MT|Und|Unidad)?\s*(.+)$', linea_str, re.IGNORECASE)
            if match_a:
                codigo = match_a.group(1).upper()
                cant_val = float(match_a.group(2).replace(',', '.'))
                desc_bruta = match_a.group(3).strip()

        if codigo and cant_val is not None and desc_bruta:
            if any(exc in codigo for exc in excluidos):
                continue

            if any(term in codigo.upper() for term in ['CLIENTE', 'CARRERA', 'CALLE']) or any(term in desc_bruta.upper() for term in ['SOCIEDAD', 'ANONIMA', 'S.A.', 'LTDA']):
                continue

            for stop_word in ['DECLARAMOS', 'ÍTEM', 'RESPONSABLE', 'GARANTÍA']:
                if stop_word in desc_bruta.upper():
                    desc_bruta = desc_bruta.split(stop_word)[0].strip()

            desc_bruta = re.sub(r'\s+\d{1,2}$', '', desc_bruta).strip()

            desc_limpia = re.sub(r'(?:UND|MTS|MT|UNDS)\s*$', '', desc_bruta, flags=re.IGNORECASE).strip()
            desc_limpia = re.sub(r'([a-záéíóúÑA-Z])(UND|MTS|MT|UNDS)\b', r'\1', desc_limpia, flags=re.IGNORECASE).strip()
            desc_limpia = re.sub(
                r'^[\s.:,\-]*\d+(?:[.,]\d+)?\s*(?:UND|MTS|MT|UNDS|UNIDAD)\b[\s.:,\-]*',
                '',
                desc_limpia,
                flags=re.IGNORECASE
            ).strip()
            desc_limpia = re.sub(r'\s+', ' ', desc_limpia).strip()

            if not any(item['Código'] == codigo for item in items):
                items.append({
                    "Entrega": nombre_doc,
                    "Código": codigo,
                    "Descripción": desc_limpia,
                    "Cantidad": cant_val
                })
                registrar_log(f"[{nombre_doc}] Ítem capturado -> Código: {codigo} | Cantidad: {cant_val} | Desc: {desc_limpia}")

    return items

# ---------------------------------------------------------
# Motor de Parseo Logístico con Corrección de Cliente y Contacto
# ---------------------------------------------------------
def extract_pdf_full_text(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text
    except Exception:
        return ""

def extract_metadata_general(text):
    meta = {
        "No. Remisión": "No detectado",
        "Fecha Emisión": "No detectado",
        "Cliente / Empresa": "No detectado",
        "Ciudad Destino": "No detectado"
    }

    if (rem_m := re.search(r"(?:No\.|N°|Remisión|Entrega)\s*[:#]?\s*(\d{6,10})", text, re.IGNORECASE)):
        meta["No. Remisión"] = rem_m.group(1).strip()

    if (f_m := re.search(r"Fecha\s*[:]?\s*(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)):
        meta["Fecha Emisión"] = f_m.group(1).strip()

    banned_headers = ["ENTREGA", "MERCANCIA", "REMISIÓN", "ORDEN", "PEDIDOS", "DIRECCIÓN"]
    for line in text.split('\n'):
        if "CLIENTE" in line.upper() and ":" in line:
            val = line.split(":", 1)[1].strip()
            if len(val) > 3 and not any(b in val.upper() for b in banned_headers):
                meta["Cliente / Empresa"] = val
                break
    if meta["Cliente / Empresa"] == "No detectado":
        if (cli_m := re.search(r"CLIENTE\s*[:]?\s*([^\n]+)", text, re.IGNORECASE)):
            val = cli_m.group(1).strip()
            if not any(b in val.upper() for b in banned_headers):
                meta["Cliente / Empresa"] = val

    if (ciu_m := re.search(r"Ciudad\s*[:]?\s*([^\n]+)", text, re.IGNORECASE)):
        meta["Ciudad Destino"] = ciu_m.group(1).strip()

    return meta

def extract_observation_text(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            first_page = pdf.pages[0]
            text = first_page.extract_text()
            
        pattern = r"Observación:\s*(.*?)(?=\n[A-Z]|$)"
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        
        if match:
            obs_raw = match.group(1).replace("\n", " ").strip()
            return re.split(r'Basado en Pedidos', obs_raw, flags=re.IGNORECASE)[0].strip()
        
        return text[:300].replace("\n", " ")
    except Exception:
        return None

def parse_observation_data(text):
    data = {
        "Dirección": "No detectado",
        "Contacto": "No detectado",
        "Celular": "No detectado",
        "OC": "No detectado"
    }

    text_clean = text.replace("\n", " ").strip()

    cel_match = re.search(r'\b(3\d{2}[\s\-]?\d{3}[\s\-]?\d{4}|\b3\d{9}\b|\b\d{10}\b)', text_clean)
    if cel_match:
        data["Celular"] = cel_match.group(1).strip()

    cont_match = re.search(r"CONTACTO[:]?\s*([A-ZÁÉÍÓÚÑ\s]+?)(?=\s*-\s*TELEFONO|\s*-\s*TEL|\s*CEL\.|\s*OC|\s*//|$)", text_clean, re.IGNORECASE)
    if cont_match:
        nombre = cont_match.group(1).strip()
        if len(nombre) > 2:
            data["Contacto"] = nombre

    dir_match = re.search(r"(?:ENTREGAR EN|DIRECCION DE ENVIO|DIRECCIÓN DE ENVÍO|DIRECCIÓN|DIRECCION)[:]?\s*(.*?)(?=\s*//|\s*CONTACTO|\s*HORARIO|\s*OC|$)", text_clean, re.IGNORECASE)
    if dir_match:
        dir_val = dir_match.group(1).strip()
        data["Dirección"] = re.split(r'Basado en Pedidos', dir_val, flags=re.IGNORECASE)[0].strip()
    else:
        data["Dirección"] = text_clean

    if data["Contacto"] != "No detectado":
        data["Contacto"] = data["Contacto"].replace("PRINCIPAL", "").replace("ALMACEN", "").strip()

    return data

def procesar_ubicacion(direccion):
    dir_upper = direccion.upper()
    mun = ["PUERTO LOPEZ", "PUERTO GAITAN", "BARRANQUILLA", "MEDELLIN", "ITAGUI", "COTA", "FUNZA", "MOSQUERA", "MADRID", "FACATATIVA", "CHIA", "CAJICA", "ZIPAQUIRA", "YUMBO", "CALI", "BOGOTA", "TOCANCIPA", "SOACHA", "PUENTE ARANDA"]
    muni_det = "NO ESPECIFICADO"
    for m in mun:
        if m in dir_upper: muni_det = m; break
    if muni_det in ["FUNZA", "MOSQUERA", "MADRID", "FACATATIVA"]: reg = "SABANA OCCIDENTE"
    elif muni_det in ["COTA", "CHIA", "CAJICA", "ZIPAQUIRA", "TOCANCIPA"]: reg = "SABANA NORTE"
    elif muni_det in ["BOGOTA", "PUENTE ARANDA"]: reg = "NORTE" if any(z in dir_upper for z in ["SUBA", "USAQUEN", "CALLE 170"]) else "SUR"
    elif muni_det in ["YUMBO", "CALI", "BARRANQUILLA", "MEDELLIN", "ITAGUI"]: reg = "SUR"
    else: reg = "POR DEFINIR"
    return muni_det, reg

def generar_link_maps(direccion):
    if not direccion or direccion == "No detectado": return "#"
    return "https://www.google.com/maps/search/?api=1&query=" + direccion.strip().replace(" ", "+")

# ---------------------------------------------------------
# Capa de Abstracción de Entregas y Renderizado
# ---------------------------------------------------------
def obtener_datos_entrega_source(num_entrega):
    env_mode = st.session_state.get("env_mode", "DEV (Google)")
    if env_mode == "PROD (SAP)":
        registrar_log(f"[SAP BTP ODATA MOCK] Consumiendo API_OUTBOUND_DELIVERY_SRV para entrega: {num_entrega}", "INFO")
        return [
            {"Entrega": f"Entrega_{num_entrega}", "Código": "45230050", "Descripción": "[SAP S/4HANA] TUBERIA PVC", "Cantidad": 10.0},
            {"Entrega": f"Entrega_{num_entrega}", "Código": "45210080", "Descripción": "[SAP S/4HANA] ACCESORIO PVC", "Cantidad": 5.0}
        ]
    else:
        pdf_buffer = descargar_pdf_desde_drive(DRIVE_FOLDER_ID, num_entrega)
        if pdf_buffer: return extraer_tabla_materiales(pdf_buffer, nombre_doc=f"Entrega_{num_entrega}")
        return []

def render_procesamiento_despacho(lista_fuentes, tab_key, mostrar_exportacion=True, registrar_func=registrar_envio_en_supabase):
    if not lista_fuentes:
        return

    fuentes_unicas = []
    tags_vistos = set()
    for fuente, doc_clean, destino, cliente, instrucciones in lista_fuentes:
        clave = (doc_clean, destino, cliente, instrucciones)
        if clave not in tags_vistos:
            tags_vistos.add(clave)
            fuentes_unicas.append((fuente, doc_clean, destino, cliente, instrucciones))

    todos_los_items = []
    for pdf_source, doc_clean, destino, cliente, instrucciones in fuentes_unicas:
        try:
            items_doc = extraer_tabla_materiales(pdf_source, nombre_doc=doc_clean)
            for item in items_doc:
                item["Destino"] = destino
                item["Cliente"] = cliente
                item["Instrucciones"] = instrucciones
            todos_los_items.extend(items_doc)
        except Exception as e:
            st.error(f"Error procesando el documento `{doc_clean}`: {e}")
            registrar_log(f"Error en documento {doc_clean}: {e}", "ERROR")

    if todos_los_items:
        df_resumen = pd.DataFrame(todos_los_items)
        pesos_u = []
        pesos_t = []

        for _, row in df_resumen.iterrows():
            codigo_item = str(row['Código']).strip()
            item_match = get_product_data_from_source(codigo_item, df_bd)

            if item_match is not None:
                p_raw = str(item_match.get('Peso_KG', 0.0)).strip().replace(',', '.').replace(' ', '')
                try:
                    p_val = float(p_raw)
                except ValueError:
                    p_val = 0.0
            else:
                p_val = 0.0

            cant_val = float(row['Cantidad'])
            pesos_u.append(p_val)
            pesos_t.append(round(p_val * cant_val, 2))

        df_resumen["Peso Unit. (KG)"] = pesos_u
        df_resumen["Peso Total (KG)"] = pesos_t
        df_resumen["No."] = [str(i + 1) for i in range(len(df_resumen))]

        columnas_orden = ["No.", "Entrega", "Código", "Descripción", "Cantidad"]
        if "Peso Unit. (KG)" in df_resumen.columns:
            columnas_orden.extend(["Peso Unit. (KG)", "Peso Total (KG)"])

        df_resumen = df_resumen[columnas_orden + ["Destino"]]

        st.markdown(f"### 📋 Tabla Consolidada de Materiales ({len(fuentes_unicas)} Documentos)")
        st.dataframe(df_resumen.drop(columns=["Destino"]), use_container_width=True)

        if "Peso Total (KG)" in df_resumen.columns:
            total_kg = pd.to_numeric(df_resumen["Peso Total (KG)"], errors='coerce').sum()
            total_ton = total_kg / 1000.0
            total_docs = len(fuentes_unicas)

            st.markdown("---")
            st.markdown("### 📊 Gran Total Consolidado del Despacho")
            m1, m2, m3 = st.columns(3)
            m1.metric("📄 Total Documentos Procesados", f"{total_docs}")
            m2.metric("📦 Peso Total Combinado (KG)", f"{total_kg:,.2f} KG")
            m3.metric("🚛 Peso Total Combinado (Toneladas)", f"{total_ton:,.3f} T")

            if mostrar_exportacion:
                st.markdown("---")
                revision_formulario = st.session_state.get(f"limpieza_{tab_key}", 0)
                is_local = tab_key == "tab_local"
                
                if f"saved_data_{tab_key}" not in st.session_state or not isinstance(st.session_state[f"saved_data_{tab_key}"], dict):
                    st.session_state[f"saved_data_{tab_key}"] = {
                        "dest_name": "DESPACHOS CLIENTE / REMISIÓN",
                        "dest_address": "DIRECCIÓN REGISTRADA EN REMISIONES",
                        "d_nombre": "",
                        "d_placa": "",
                        "d_cedula": "",
                        "d_marca": "",
                        "d_celular": "",
                        "d_transp": "",
                        "elab_nombre": "LOGÍSTICA TUVACOL S.A.",
                        "empaques_info": "Ninguno especificado"
                    }

                with st.form(key=f"form_despacho_{tab_key}"):
                    if not is_local:
                        st.markdown("### 📄 Datos del Destinatario")
                        col_dst1, col_dst2 = st.columns(2)
                        with col_dst1:
                            dest_name_input = st.text_input("Nombre del Destinatario:", value=st.session_state[f"saved_data_{tab_key}"]["dest_name"])
                        with col_dst2:
                            dest_address_input = st.text_input("Dirección / Teléfono del Destinatario:", value=st.session_state[f"saved_data_{tab_key}"]["dest_address"])
                    else:
                        dest_name_input = "LOCAL"
                        dest_address_input = "LOCAL"

                    st.markdown("---")
                    st.markdown("### 🚛 Datos del Conductor y Vehiculo")
                    if not is_local:
                        col_drv1, col_drv2, col_drv3 = st.columns(3)
                        with col_drv1:
                            d_nombre_input = st.text_input("Nombre Completo Conductor:", value=st.session_state[f"saved_data_{tab_key}"]["d_nombre"])
                            d_placa_input = st.text_input("Placa Vehículo:", value=st.session_state[f"saved_data_{tab_key}"]["d_placa"])
                        with col_drv2:
                            d_cedula_input = st.text_input("Cédula No.:", value=st.session_state[f"saved_data_{tab_key}"]["d_cedula"])
                            d_marca_input = st.text_input("Marca / Modelo / Color:", value=st.session_state[f"saved_data_{tab_key}"]["d_marca"])
                        with col_drv3:
                            d_celular_input = st.text_input("Celular / Teléfono:", value=st.session_state[f"saved_data_{tab_key}"]["d_celular"])
                            d_transp_input = st.text_input("Empresa Transportadora:", value=st.session_state[f"saved_data_{tab_key}"]["d_transp"])
                    else:
                        col_drv1, col_drv2 = st.columns(2)
                        with col_drv1:
                            d_nombre_input = st.text_input("Nombre Completo Conductor:", value=st.session_state[f"saved_data_{tab_key}"]["d_nombre"])
                            d_cedula_input = st.text_input("Cédula No.:", value=st.session_state[f"saved_data_{tab_key}"]["d_cedula"])
                        with col_drv2:
                            d_placa_input = st.text_input("Placa Vehículo:", value=st.session_state[f"saved_data_{tab_key}"]["d_placa"])
                        d_marca_input = ""
                        d_celular_input = ""
                        d_transp_input = ""

                    st.markdown("---")
                    st.markdown("### 📦 Cantidad de Empaques por Tipo")
                    st.caption("Indique la cantidad utilizada para cada tipo de empaque en este despacho.")
                    empaque_cols1 = st.columns(3)
                    with empaque_cols1[0]:
                        cant_estibas = st.number_input("Estibas", min_value=0, value=0, step=1, key=f"estiba_{tab_key}_{revision_formulario}")
                    with empaque_cols1[1]:
                        cant_guacales = st.number_input("Guacales", min_value=0, value=0, step=1, key=f"guacal_{tab_key}_{revision_formulario}")
                    with empaque_cols1[2]:
                        cant_cajas = st.number_input("Cajas", min_value=0, value=0, step=1, key=f"caja_{tab_key}_{revision_formulario}")

                    empaque_cols2 = st.columns(3)
                    with empaque_cols2[0]:
                        cant_sobres = st.number_input("Sobres", min_value=0, value=0, step=1, key=f"sobre_{tab_key}_{revision_formulario}")
                    with empaque_cols2[1]:
                        cant_paquetes = st.number_input("Paquetes", min_value=0, value=0, step=1, key=f"paquete_{tab_key}_{revision_formulario}")
                    with empaque_cols2[2]:
                        cant_tubos = st.number_input("Tubos", min_value=0, value=0, step=1, key=f"tubo_{tab_key}_{revision_formulario}")

                    resumen_empaques = []
                    if cant_estibas > 0: resumen_empaques.append(f"{cant_estibas} Estiba(s)")
                    if cant_guacales > 0: resumen_empaques.append(f"{cant_guacales} Guacal(es)")
                    if cant_cajas > 0: resumen_empaques.append(f"{cant_cajas} Caja(s)")
                    if cant_sobres > 0: resumen_empaques.append(f"{cant_sobres} Sobre(s)")
                    if cant_paquetes > 0: resumen_empaques.append(f"{cant_paquetes} Paquete(s)")
                    if cant_tubos > 0: resumen_empaques.append(f"{cant_tubos} Tubo(s)")
                    empaques_input = ", ".join(resumen_empaques) if resumen_empaques else "Ninguno especificado"

                    st.markdown("---")
                    st.markdown("### ✍️ Información de Elaboración")
                    elab_nombre_input = st.text_input("Elaborado por (Nombre y Cargo):", value=st.session_state[f"saved_data_{tab_key}"]["elab_nombre"])

                    submitted = st.form_submit_button(label="💾 Guardar Cambios")

                    if submitted:
                        campos_obligatorios = [d_nombre_input, d_cedula_input, d_placa_input, elab_nombre_input]
                        if not is_local:
                            campos_obligatorios.extend([
                                dest_name_input, dest_address_input, d_marca_input,
                                d_celular_input, d_transp_input
                            ])
                        if not all(campo.strip() for campo in campos_obligatorios):
                            st.session_state[f"saved_data_{tab_key}"] = None
                            st.session_state.pop(f"consecutivo_num_{tab_key}", None)
                            st.error("❌ Faltan campos obligatorios por llenar. Complete los datos del destinatario, conductor, vehículo y elaboración antes de guardar.")
                        else:
                            st.session_state[f"saved_data_{tab_key}"] = {
                                "dest_name": dest_name_input.strip(),
                                "dest_address": dest_address_input.strip(),
                                "d_nombre": d_nombre_input.strip(),
                                "d_placa": d_placa_input.strip(),
                                "d_cedula": d_cedula_input.strip(),
                                "d_marca": d_marca_input.strip(),
                                "d_celular": d_celular_input.strip(),
                                "d_transp": d_transp_input.strip(),
                                "elab_nombre": elab_nombre_input.strip(),
                                "empaques_info": empaques_input
                            }
                            observaciones_guardado = (
                                f"TOTAL: {df_resumen['Cantidad'].sum():,.0f} UND, "
                                f"{total_kg:,.2f} KG | DOCS: {', '.join(sorted(set(df_resumen['Entrega'].astype(str))))} | "
                                f"Empaques: {empaques_input}"
                            )
                            consecutivo_num = registrar_func(
                                st.session_state[f"saved_data_{tab_key}"],
                                df_resumen,
                                observaciones_guardado
                            )
                            if consecutivo_num is None:
                                st.session_state[f"saved_data_{tab_key}"] = None
                                st.error("❌ No fue posible registrar el envío en Supabase. Los formatos siguen bloqueados.")
                            else:
                                st.session_state[f"consecutivo_num_{tab_key}"] = consecutivo_num
                                st.success(f"✅ Datos guardados. Consecutivo asignado: No.{str(consecutivo_num).zfill(8)}")

                saved = st.session_state[f"saved_data_{tab_key}"]
                st.markdown("---")
                st.markdown("### Generar Relación De Envio De Mercancia")
                if saved is None or st.session_state.get(f"consecutivo_num_{tab_key}") is None:
                    st.warning("🔒 Complete todos los campos obligatorios y haga clic en **'Guardar Cambios'** para habilitar la previsualización y las descargas.")
                else:
                    dest_info = {"nombre": saved["dest_name"], "direccion": saved["dest_address"]}
                    driver_info = {
                        "nombre": saved["d_nombre"], "cedula": saved["d_cedula"], "celular": saved["d_celular"],
                        "placa": saved["d_placa"], "marca": saved["d_marca"], "transportadora": saved["d_transp"]
                    }
                    elaborado_info = {"nombre": saved["elab_nombre"]}
                    empaques_info = saved.get("empaques_info", "Ninguno especificado")

                    consecutivo_num = st.session_state.get(f"consecutivo_num_{tab_key}")
                    excel_bytes = generar_excel_bytes(df_resumen, total_docs, total_kg, total_ton, driver_info, dest_info, elaborado_info, empaques_info, consecutivo_num)
                    observaciones_pdf = f"Empaques: {empaques_info}"
                    if is_local:
                        pdf_bytes = generar_pdf_local_bytes(df_resumen, saved, observaciones_pdf, consecutivo_num)
                    else:
                        pdf_bytes = generar_pdf_bytes(df_resumen, total_docs, total_kg, total_ton, driver_info, dest_info, elaborado_info, empaques_info, consecutivo_num)

                    with st.expander("👁️ Previsualizar Documento Generado (PDF)", expanded=False):
                        base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                        pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600px" type="application/pdf"></iframe>'
                        st.markdown(pdf_display, unsafe_allow_html=True)

                    registro_mensaje = st.session_state.pop("registro_descarga_mensaje", None)
                    if registro_mensaje:
                        st.info(registro_mensaje)

                    col_exp1, col_exp2 = st.columns(2)
                    with col_exp1:
                        st.download_button(
                            label="📊 Descargar Formato Oficial Excel (.xlsx)",
                            data=excel_bytes,
                            file_name=f"Relacion_Envio_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key=f"dl_excel_{tab_key}"
                        )
                    with col_exp2:
                        st.download_button(
                            label="🖨️ Descargar Formato Oficial PDF (Imprimible)",
                            data=pdf_bytes,
                            file_name=f"Relacion_Envio_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            use_container_width=True, key=f"dl_pdf_{tab_key}"
                        )
            return df_resumen
    else:
        st.warning("⚠️ No se encontraron productos válidos en los documentos seleccionados.")
    return None

# ---------------------------------------------------------
# Interfaz Principal y Pestañas
# ---------------------------------------------------------
st.sidebar.markdown(f"👤 **Usuario:** {st.session_state.usuario_actual}")
st.sidebar.markdown(f"🏷️ **Rol:** {st.session_state.get('rol_actual', 'LOGISTICA')}")

if st.sidebar.button("Cerrar Sesión", use_container_width=True):
    st.session_state.autenticado = False
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.header("🔐 Modo de Operación")
modo_app = st.sidebar.radio("Seleccione la interfaz:", ["Modo Usuario", "Modo Destroller"])

es_dev_autenticado = False
if modo_app == "Modo Destroller":
    cedula_input = st.sidebar.text_input("Ingrese Contraseña de Destroller:", type="password")
    if cedula_input.strip() == CEDULA_DEV_CORRECTA:
        st.sidebar.success("🔓 Acceso Destroller Autorizado")
        es_dev_autenticado = True

if es_dev_autenticado:
    env_mode = st.sidebar.radio("Modo Conexión de Datos:", ["DEV (Google)", "PROD (SAP)"], key="env_mode")

df_bd = fetch_google_sheet_database()
if df_bd is None: df_bd = cargar_bd_local(BD_LOCAL_PATH)

if st.sidebar.button("🔄 Sincronizar Google Sheets Oficial"):
    st.cache_data.clear()
    st.rerun()

st.title("🔄 FlowShift")
st.markdown("<p style='color: #666; font-size: 16px; margin-top: -15px;'><i>La evolución inteligente de tu gestión administrativa.</i></p>", unsafe_allow_html=True)

rol_usuario = st.session_state.get("rol_actual", "LOGISTICA")
tab1 = tab2 = tab3 = tab_local = tab_consulta = tab4 = None

if es_dev_autenticado:
    tab1, tab2, tab3, tab_local, tab_consulta, tab4 = st.tabs([
        "🔍 Búsqueda por Código", "📄 Procesar Remisión / PDF", "📤 Relación de Envío", "📤 Relación de Envío Local", "🔎 Consultar Documento", "📜 Logs"
    ])
elif "VENTAS" in rol_usuario:
    tab1, tab2 = st.tabs([
        "🔍 Búsqueda por Código", "📄 Procesar Remisión / PDF"
    ])
else:
    tab3, tab_local, tab_consulta = st.tabs([
        "📤 Relación de Envío", "📤 Relación de Envío Local", "🔎 Consultar Documento"
    ])

if tab1 is not None:
    with tab1:
        st.subheader("Consulta Dinámica de Producto")
        codigo_input = st.text_input("Ingrese Código o Descripción del Artículo", value="", placeholder="Ej. 108001051")
        cant_input = st.number_input("Cantidad a despachar", min_value=1.0, value=1.0, step=1.0)
        if codigo_input.strip() != "":
            item = get_product_data_from_source(codigo_input, df_bd)
            if item is not None:
                peso_unit = float(str(item.get('Peso_KG', 0.0)).replace(',', '.'))
                st.markdown(
                    f"""
                    <div class="product-card">
                        <p><strong>📦 Código:</strong> {item['Codigo']}</p>
                        <p><strong>📝 Descripción:</strong> {item['Descripcion']}</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                res_col1, res_col2 = st.columns(2)
                res_col1.metric("Peso Unitario", f"{peso_unit:.2f} KG")
                res_col2.metric("Peso Total", f"{peso_unit * cant_input:.2f} KG")
            else:
                st.warning("⚠️ No se encontraron coincidencias.")

if tab2 is not None:
    with tab2:
        st.subheader("Procesar Remisión / Consulta de Pesos")
        st.info("ℹ️ Esta pestaña permite consultar pesos desde archivos PDF, sin generar reportes.")
        if "limpieza_tab2" not in st.session_state:
            st.session_state.limpieza_tab2 = 0
        uploaded_files_tab2 = st.file_uploader(
            "Cargar archivos PDF de remisión",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"uploader_consulta_pesos_{st.session_state.limpieza_tab2}"
        )
        if st.button("🗑️ Limpiar PDFs y consulta", key="limpiar_tab2"):
            st.session_state.limpieza_tab2 += 1
            st.rerun()
        lista_fuentes_tab2 = []
        if uploaded_files_tab2:
            for file in uploaded_files_tab2:
                doc_clean, destino, cliente, instrucciones = analizar_metadatos_documento(file)
                lista_fuentes_tab2.append((file, doc_clean, destino, cliente, instrucciones))
        render_procesamiento_despacho(lista_fuentes_tab2, "tab2", mostrar_exportacion=False)

if tab3 is not None:
    with tab3:
        st.subheader("📤 Relación de Envío y Control de Empaques")
        if "limpieza_tab3" not in st.session_state:
            st.session_state.limpieza_tab3 = 0

        uploaded_files = st.file_uploader(
            "Cargar PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"uploader_relacion_envio_{st.session_state.limpieza_tab3}"
        )
        if st.button("🗑️ Limpiar PDFs y campos", key="limpiar_tab3"):
            st.session_state.limpieza_tab3 += 1
            st.session_state.pop("saved_data_tab3", None)
            st.session_state.pop("empaques_info_tab3", None)
            st.session_state.pop("consecutivo_num_tab3", None)
            st.rerun()

        lista_fuentes = []
        if uploaded_files:
            for file in uploaded_files:
                doc_clean, destino, cliente, instrucciones = analizar_metadatos_documento(file)
                lista_fuentes.append((file, doc_clean, destino, cliente, instrucciones))
        df_resultado_t3 = render_procesamiento_despacho(lista_fuentes, "tab3", mostrar_exportacion=True)

if tab_local is not None:
    with tab_local:
        st.subheader("📤 Relación de Envío Local y Generación de Formato Oficial")
        if "limpieza_tab_local" not in st.session_state:
            st.session_state.limpieza_tab_local = 0

        uploaded_files_local = st.file_uploader(
            "Cargar PDFs para Relación de Envío Local",
            type=["pdf"],
            accept_multiple_files=True,
            key=f"uploader_relacion_envio_local_{st.session_state.limpieza_tab_local}"
        )
        lista_fuentes_local = []
        if uploaded_files_local:
            for file in uploaded_files_local:
                doc_clean, destino, cliente, instrucciones = analizar_metadatos_documento(file)
                lista_fuentes_local.append((file, doc_clean, destino, cliente, instrucciones))
        render_procesamiento_despacho(
            lista_fuentes_local,
            "tab_local",
            mostrar_exportacion=True,
            registrar_func=registrar_envio_local_en_supabase
        )
        if st.button("🗑️ Limpiar PDFs y campos locales", key="limpiar_tab_local"):
            st.session_state.limpieza_tab_local += 1
            st.session_state.pop("saved_data_tab_local", None)
            st.session_state.pop("consecutivo_num_tab_local", None)
            st.rerun()

if tab_consulta is not None:
    with tab_consulta:
        st.subheader("🔎 Consultar Documento Original")
        st.caption("Busque una Transferencia de Stock (TS), Entrega de Mercancía (EDM) o referencia guardada en el historial.")
        busqueda_doc = st.text_input("Número de documento o referencia", placeholder="Ej. 4996 o 20006855", key="busqueda_documento_historial")

        if st.button("🔍 Buscar en Historial", key="btn_buscar_historial"):
            termino = busqueda_doc.strip()
            if not termino:
                st.warning("⚠️ Ingrese un número de documento antes de buscar.")
            else:
                try:
                    response_general = (
                        supabase.table("historial_envios")
                        .select("*")
                        .ilike("observaciones", f"%{termino}%")
                        .execute()
                    )
                    response_local = (
                        supabase.table("historial_envios_local")
                        .select("*")
                        .ilike("observaciones", f"%{termino}%")
                        .execute()
                    )
                    registros = (response_general.data or []) + (response_local.data or [])
                    registros.sort(key=lambda registro: registro.get("consecutivo", registro.get("id", 0)) or 0, reverse=True)
                    if not registros:
                        st.warning(f"⚠️ No se encontró ningún despacho asociado a '{termino}'.")
                    else:
                        st.success(f"✅ Se encontraron {len(registros)} despacho(s) asociado(s) a '{termino}'.")
                        for registro in registros:
                            consecutivo = registro.get("consecutivo", registro.get("id", "N/A"))
                            fecha = formatear_fecha_colombia(registro.get("fecha") or registro.get("created_at"))
                            st.markdown(
                                f"""
                                <div class="product-card">
                                    <h4 style="color: #1F3864; margin-top: 0;">Datos del vehículo y conductor</h4>
                                    <p><b>Conductor:</b> {registro.get('conductor_nombre', 'N/A')}</p>
                                    <p><b>Cédula:</b> {registro.get('conductor_cedula', 'N/A')}</p>
                                    <p><b>Placa:</b> {registro.get('vehiculo_placa', 'N/A')}</p>
                                    <p><b>Fecha:</b> {fecha}</p>
                                    <p><b>Peso total:</b> {float(registro.get('peso_total_kg') or 0):,.2f} KG</p>
                                    <p><b>Documentos:</b> {registro.get('total_documentos', 0)}</p>
                                </div>
                                """,
                                unsafe_allow_html=True
                            )

                            items_json = registro.get("items_json")
                            saved_data_json = registro.get("saved_data_json")
                            if items_json and saved_data_json:
                                try:
                                    df_historial = pd.DataFrame(items_json)
                                    pdf_historial = generar_pdf_bytes(
                                        df_historial,
                                        len(set(df_historial["Entrega"])),
                                        float(df_historial["Peso Total (KG)"].sum()),
                                        float(df_historial["Peso Total (KG)"].sum()) / 1000,
                                        {
                                            "nombre": saved_data_json.get("d_nombre", ""),
                                            "cedula": saved_data_json.get("d_cedula", ""),
                                            "celular": saved_data_json.get("d_celular", ""),
                                            "placa": saved_data_json.get("d_placa", ""),
                                            "marca": saved_data_json.get("d_marca", ""),
                                            "transportadora": saved_data_json.get("d_transp", "")
                                        },
                                        {
                                            "nombre": saved_data_json.get("dest_name", ""),
                                            "direccion": saved_data_json.get("dest_address", "")
                                        },
                                        {"nombre": saved_data_json.get("elab_nombre", "")},
                                        saved_data_json.get("empaques_info", "Ninguno especificado"),
                                        consecutivo,
                                        es_local=not bool(saved_data_json.get("cond_empresa"))
                                    )
                                    st.download_button(
                                        label=f"📄 Descargar PDF Oficial (No.{str(consecutivo).zfill(8)})",
                                        data=pdf_historial,
                                        file_name=f"Relacion_Envio_No_{str(consecutivo).zfill(8)}.pdf",
                                        mime="application/pdf",
                                        key=f"dl_historial_{registro.get('id', consecutivo)}"
                                    )
                                except Exception as error_pdf:
                                    registrar_log(f"Error generando PDF histórico: {error_pdf}", "WARNING")
                                    st.caption("ℹ️ No fue posible reconstruir el PDF de este registro.")
                            else:
                                st.caption("ℹ️ Este registro anterior no tiene el detalle necesario para descargar el PDF.")
                except Exception as e:
                    registrar_log(f"Error consultando historial por documento: {e}", "ERROR")
                    st.error(f"❌ Error al consultar la base de datos: {e}")

if es_dev_autenticado and tab4 is not None:
    with tab4:
        st.subheader("📜 Log Auditoría de Ejecución")
        if os.path.exists(LOG_FILENAME):
            try:
                with open(LOG_FILENAME, "r", encoding="utf-8") as f:
                    log_content = f.read()
                if log_content.strip():
                    st.download_button("💾 Descargar log", log_content, LOG_FILENAME, "text/plain")
                    st.text_area("Contenido del Log:", value=log_content, height=400)
                else:
                    st.info("ℹ️ El archivo de log está vacío por el momento.")
            except Exception as e:
                st.error(f"Error leyendo el archivo de log: {e}")
        else:
            st.warning("⚠️ El archivo 'log_ejecucion.txt' aún no ha sido creado.")
